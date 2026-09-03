#!/usr/bin/env python3
import os
import sys
import math
import time
import shutil
import zipfile
import ctypes
import threading
import csv
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

import polars as pl
import xlsxwriter
import openpyxl  # Soporte para lectura e inspección de estructuras Excel
from PyPDF2 import PdfReader, PdfWriter
import pikepdf

import pymupdf
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


# ==============================================================================
# CACHÉ Y DETECTOR ROBUSTO DE ENCABEZADOS DE ARCHIVOS (CSV & EXCEL)
# ==============================================================================

_CLEAN_FILES_CACHE = {}

def es_encabezado_real(campos):
    """
    Detecta el encabezado buscando la transición de 'predominantemente texto' 
    a 'predominantemente datos numéricos'.
    """
    if not campos or len(campos) < 2:
        return False

    # 1. Analizamos la composición de la fila
    total_campos = len(campos)
    textos = 0
    numeros = 0
    vacios = 0

    for c in campos:
        val = str(c).strip()
        if not val:
            vacios += 1
        elif val.replace('.', '', 1).replace('-', '', 1).isdigit():
            numeros += 1
        else:
            textos += 1

    # 2. Un encabezado real tiene:
    # - Predominancia de texto (nombres de columna)
    # - Muy pocos números (no es una fila de datos)
    
    es_casi_todo_texto = (textos / total_campos) > 0.6  # Más del 60% es texto
    hay_pocos_numeros = (numeros / total_campos) < 0.2  # Menos del 20% son números
    
    return es_casi_todo_texto and hay_pocos_numeros

def obtener_archivo_limpio(filepath):
    """
    Identifica el encabezado real en CSV o Excel (.xlsx, .xlsm), omitiendo títulos
    o texto no relevante superior, y retorna una ruta de trabajo estandarizada.
    """
    if filepath in _CLEAN_FILES_CACHE and os.path.exists(_CLEAN_FILES_CACHE[filepath][0]):
        return _CLEAN_FILES_CACHE[filepath]

    ext = os.path.splitext(filepath)[1].lower()

    # --- PROCESAMIENTO PARA ARCHIVOS EXCEL ---
    if ext in ['.xlsx', '.xlsm', '.xls']:
        try:
            temp_dir = os.path.join(os.path.dirname(filepath), "_temp_cleaned_csv")
            os.makedirs(temp_dir, exist_ok=True)
            clean_filename = f"clean_xl_{os.path.basename(filepath)}.csv"
            clean_path = os.path.join(temp_dir, clean_filename)

            header_escrito = False

            # Usar utf-8-sig (con BOM) para garantizar la correcta interpretación de caracteres especiales
            with open(clean_path, 'w', encoding='utf-8-sig', newline='') as f_out:
                writer = csv.writer(f_out)

                if ext == '.xls':
                    wb = xlrd.open_workbook(filepath)
                    hojas = [wb.sheet_by_index(i) for i in range(wb.nsheets)]
                else:
                    # Sin data_only=True para evitar que openpyxl convierta enteros en flotantes automáticamente
                    wb = openpyxl.load_workbook(filepath, read_only=True)
                    hojas = wb.worksheets

                for sheet in hojas:
                    if ext == '.xls':
                        filas = [sheet.row_values(r) for r in range(sheet.nrows)]
                    else:
                        filas = list(sheet.iter_rows(values_only=True))

                    # --- Normalizar longitud de filas ---
                    # En modo read_only, openpyxl puede recortar celdas finales vacías
                    # de forma independiente por fila cuando el archivo (típico de
                    # exportes de sistemas contables como QuickBooks) no trae bien
                    # declarado el rango <dimension> interno. Esto produce filas de
                    # distinto largo ("dentadas"), lo que luego rompe pl.scan_csv con
                    # "found more fields than defined in Schema". Se rellenan con None
                    # todas las filas hasta el largo máximo detectado en la hoja.
                    if filas:
                        max_len = max(len(r) for r in filas)
                        filas = [
                            tuple(r) + (None,) * (max_len - len(r)) if len(r) < max_len else r
                            for r in filas
                        ]

                    header_idx = None
                    for idx, row in enumerate(filas[:150]):
                        if not row: continue
                        row_vals = [str(c).strip() if c is not None else "" for c in row]
                        if any(row_vals) and es_encabezado_real(row_vals):
                            header_idx = idx
                            break

                    if header_idx is None:
                        continue  # Pestaña vacía o sin estructura válida

                    # Escribir contenido formateando números enteros para que no tengan ".0"
                    for idx, row in enumerate(filas):
                        if idx < header_idx:
                            continue  # Salta metadatos superiores
                        if idx == header_idx and header_escrito:
                            continue  # Evita duplicar el encabezado

                        # Limpieza de valores: Si es flotante entero (ej. 4158.0), convertir a "4158"
                        fila_limpia = []
                        for c in row:
                            if c is None:
                                fila_limpia.append("")
                            elif isinstance(c, float) and c.is_integer():
                                fila_limpia.append(str(int(c)))
                            else:
                                fila_limpia.append(str(c))

                        writer.writerow(fila_limpia)

                    header_escrito = True

                if ext != '.xls': wb.close()

            _CLEAN_FILES_CACHE[filepath] = (clean_path, ',')
            return clean_path, ','
        except Exception:
            _CLEAN_FILES_CACHE[filepath] = (filepath, ',')
            return filepath, ','

    # --- PROCESAMIENTO PARA ARCHIVOS CSV ---
    encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
    for enc in encodings:
        try:
            lineas_raw = []
            with open(filepath, 'r', encoding=enc) as f:
                for _ in range(150):
                    line = f.readline()
                    if not line:
                        break
                    lineas_raw.append(line)

            if not lineas_raw:
                continue

            separadores = [',', ';', '\t', '|']
            max_cols_sep = {}

            for s in separadores:
                max_c = 0
                for l in lineas_raw:
                    if not l.strip():
                        continue
                    try:
                        c = next(csv.reader([l], delimiter=s), [])
                        if len(c) > max_c:
                            max_c = len(c)
                    except Exception:
                        c = l.split(s)
                        if len(c) > max_c:
                            max_c = len(c)
                max_cols_sep[s] = max_c

            best_sep = max(max_cols_sep, key=max_cols_sep.get)
            max_cols = max_cols_sep[best_sep]
            if max_cols <= 1:
                best_sep = ','

            target_header_row = None
            header_record_index = -1

            with open(filepath, 'r', encoding=enc) as f_obj:
                reader = csv.reader(f_obj, delimiter=best_sep)
                for idx, row in enumerate(reader):
                    if idx > 150:
                        break
                    if len(row) >= max_cols - 1 and len(row) > 1:
                        if es_encabezado_real(row):
                            target_header_row = row
                            header_record_index = idx
                            break

            if header_record_index <= 0 or not target_header_row:
                _CLEAN_FILES_CACHE[filepath] = (filepath, best_sep)
                return filepath, best_sep

            col_1 = target_header_row[0].strip()
            col_2 = target_header_row[1].strip() if len(target_header_row) > 1 else ""

            linea_fisica_idx = -1
            with open(filepath, 'r', encoding=enc) as f:
                for l_idx, l in enumerate(f):
                    if col_1 in l and (not col_2 or col_2 in l):
                        linea_fisica_idx = l_idx
                        break

            if linea_fisica_idx <= 0:
                _CLEAN_FILES_CACHE[filepath] = (filepath, best_sep)
                return filepath, best_sep

            temp_dir = os.path.join(os.path.dirname(filepath), "_temp_cleaned_csv")
            os.makedirs(temp_dir, exist_ok=True)
            clean_filename = f"clean_{os.path.basename(filepath)}"
            clean_path = os.path.join(temp_dir, clean_filename)

            # Escribir con utf-8-sig para preservar el símbolo "·" sin corrupción "Â·"
            with open(filepath, 'r', encoding=enc) as f_in:
                for _ in range(linea_fisica_idx):
                    f_in.readline()
                with open(clean_path, 'w', encoding='utf-8-sig', newline='') as f_out:
                    shutil.copyfileobj(f_in, f_out)

            _CLEAN_FILES_CACHE[filepath] = (clean_path, best_sep)
            return clean_path, best_sep

        except Exception:
            continue

    _CLEAN_FILES_CACHE[filepath] = (filepath, ',')
    return filepath, ','


# ==============================================================================
# FUNCIONES AUXILIARES DE PDF
# ==============================================================================

def crear_pagina_indice_pdf(items, desfase_paginas, archivo_temp_indice):
    doc = SimpleDocTemplate(
        archivo_temp_indice,
        pagesize=letter,
        leftMargin=54, rightMargin=54,
        topMargin=54, bottomMargin=54
    )
    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        'TituloIndice', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=20, leading=24,
        alignment=1, textColor=colors.HexColor("#1A202C"), spaceAfter=20
    )
    estilo_texto = ParagraphStyle(
        'TextoCapitulo', fontName='Helvetica-Bold',
        fontSize=11, leading=14, textColor=colors.HexColor("#2D3748")
    )
    estilo_pagina = ParagraphStyle(
        'TextoPagina', fontName='Helvetica-Bold',
        fontSize=11, leading=14, alignment=2, textColor=colors.HexColor("#2D3748")
    )
    
    elementos = [Paragraph("ÍNDICE GENERAL", estilo_titulo), Spacer(1, 15)]
    datos_tabla = []
    
    for pag_orig, titulo in items:
        pag_impresa = pag_orig + desfase_paginas
        datos_tabla.append([Paragraph(titulo, estilo_texto), Paragraph(str(pag_impresa), estilo_pagina)])

    tabla = Table(datos_tabla, colWidths=[400, 100])
    tabla.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
    ]))

    elementos.append(tabla)
    doc.build(elementos)


def ensamblar_y_vincular_pdf(pdf_entrada, items, pdf_indice_temp, pdf_salida, desfase, log_func=print):
    doc_final = pymupdf.open()
    doc_indice = pymupdf.open(pdf_indice_temp)
    doc_original = pymupdf.open(pdf_entrada)

    doc_final.insert_pdf(doc_indice)
    doc_final.insert_pdf(doc_original)

    toc = [[1, "Índice General", 1]]
    vinculos_creados = 0

    log_func("  🔗 Insertando hipervínculos navegables en el índice...")
    for pag_orig, titulo in items:
        pag_destino = pag_orig + desfase
        encontrado = False
        
        for p_idx in range(desfase):
            pagina_indice = doc_final[p_idx]
            rects = pagina_indice.search_for(titulo)
            
            if rects:
                y0, y1 = rects[0].y0, rects[0].y1
                rect_enlace = pymupdf.Rect(54, y0 - 5, pagina_indice.rect.width - 54, y1 + 5)
                
                link_dict = {
                    "kind": pymupdf.LINK_GOTO,
                    "page": pag_destino - 1, 
                    "from": rect_enlace
                }
                pagina_indice.insert_link(link_dict)
                vinculos_creados += 1
                encontrado = True
                break 
        
        if encontrado:
            log_func(f"     ✔ Vinculado: '{titulo}' ➔ Pág. destino {pag_destino}")
        else:
            log_func(f"     ⚠️ No se pudo vincular visualmente: '{titulo}'")
            
        toc.append([1, titulo, pag_destino])

    log_func(f"  📌 Total de hipervínculos insertados: {vinculos_creados}")
    log_func("  🔖 Generando estructura de marcadores (Bookmarks / TOC)...")
    doc_final.set_toc(toc)
    
    log_func("  💾 Guardando documento final comprimido...")
    doc_final.save(pdf_salida, garbage=4, deflate=True)
    
    doc_final.close()
    doc_indice.close()
    doc_original.close()


# ==============================================================================
# FUNCIÓN DE FORMATEO Y SANEAMIENTO UNIVERSAL DE COLUMNAS (AGNÓSTICA)
# ==============================================================================

def es_columna_numerica_disfrazada(s: pl.Series, muestra_n: int = 100) -> bool:
    """
    Evalúa si una columna de texto contiene principalmente valores numéricos
    después de remover comas de miles, comillas, espacios y símbolos monetarios.
    """
    validos = s.filter(s.is_not_null() & (s.str.strip_chars() != "")).head(muestra_n)
    if len(validos) == 0:
        return False

    # Quitamos comas de miles, comillas, espacios y caracteres no numéricos comunes
    limpios = validos.str.replace_all(r'[,\s"]', "").str.replace_all(r"[^\d.-]", "")
    
    # Expresión regular que valida formato numérico (enteros o decimales)
    patron_numero = r"^-?\d+(\.\d+)?$"
    coincidencias = limpios.str.contains(patron_numero).sum()
    
    # Si al menos el 60% de la muestra son números, la columna se considera numérica
    return (coincidencias / len(validos)) >= 0.6


def aplicar_formato_limpio(df: pl.DataFrame) -> pl.DataFrame:
    """
    Normaliza y limpia cualquier DataFrame de forma 100% agnóstica a los 
    nombres de columnas o al origen del archivo.
    """
    exprs = []
    
    for col in df.columns:
        dtype = df.schema[col]
        
        # 1. Numéricos declarados nativamente (Floats)
        if dtype in (pl.Float64, pl.Float32):
            exprs.append(
                pl.col(col).fill_null(0.0).round(4).alias(col)
            )
            
        # 2. Enteros declarados nativamente
        elif dtype in (pl.Int64, pl.Int32, pl.Int16, pl.Int8, pl.UInt64, pl.UInt32, pl.UInt16, pl.UInt8):
            exprs.append(
                pl.col(col).fill_null(0).alias(col)
            )
            
        # 3. Columnas tipo Texto / Cadenas (Evalúa si son números entrecomillados o con comas)
        else:
            serie_col = df[col]
            
            if es_columna_numerica_disfrazada(serie_col):
                # Elimina comas de miles, comillas dobles, espacios y símbolos monetarios
                exprs.append(
                    pl.col(col)
                    .cast(pl.Utf8)
                    .str.replace_all(r'[,\s"]', "")       # Remueve comas de miles y comillas
                    .str.replace_all(r"[^\d.-]", "")      # Remueve otros caracteres no numéricos
                    .replace("", None)
                    .cast(pl.Float64, strict=False)       # Casteo seguro a Float64
                    .fill_null(0.0)
                    .round(4)
                    .alias(col)
                )
            else:
                # Texto general de la tabla (Nombres, CAI, RTN, Conceptos)
                exprs.append(
                    pl.col(col)
                    .cast(pl.Utf8)
                    .str.replace(r"\s+00:00:00.*$", "")
                    .str.replace(r"\.0$", "")
                    .str.strip_chars()
                    .alias(col)
                )
            
    return df.with_columns(exprs)


def aplicar_formato_personalizado(df: pl.DataFrame, tipos_columnas: dict) -> pl.DataFrame:
    """
    Aplica los tipos de datos seleccionados manualmente por el usuario.
    """
    exprs = []
    
    for col in df.columns:
        tipo = tipos_columnas.get(col, "Texto")
        
        if tipo == "Float (Decimal)":
            exprs.append(
                pl.col(col)
                .cast(pl.Utf8)
                .str.replace_all(r'[,\s"]', "")
                .str.replace_all(r"[^\d.-]", "")
                .replace("", None)
                .cast(pl.Float64, strict=False)
                .fill_null(0.0)
                .round(4)
                .alias(col)
            )
        elif tipo == "Entero (Int)":
            exprs.append(
                pl.col(col)
                .cast(pl.Utf8)
                .str.replace_all(r'[,\s"]', "")
                .str.replace_all(r"[^\d-]", "")
                .replace("", None)
                .cast(pl.Int64, strict=False)
                .fill_null(0)
                .alias(col)
            )
        elif tipo == "Fecha":
            exprs.append(
                pl.col(col)
                .cast(pl.Utf8)
                .str.slice(0, 10)
                .alias(col)
            )
        else:  # "Texto"
            exprs.append(
                pl.col(col)
                .cast(pl.Utf8)
                .str.replace(r"\s+00:00:00.*$", "")
                .str.strip_chars()
                .fill_null("")
                .alias(col)
            )
            
    return df.with_columns(exprs)
    
# ==============================================================================
# CLASE PRINCIPAL DE LA SUITE
# ==============================================================================

class SuiteContableIntegrada:
    def __init__(self, root):
        self.root = root
        self.root.title("Suite de Análisis de Datos y Documentos (CSV / Excel / Parquet / PDF)")
        self.root.geometry("780x780")
        self.root.minsize(700, 680)

        self.paths_sep = []
        self.paths_res = []
        self.paths_conv = []

        self.pdf_path = tk.StringVar()
        self.output_dir_pdf = tk.StringVar()
        self.max_size_mb = tk.DoubleVar(value=14.0)
        self.export_mode_pdf = tk.StringVar(value="both")
        self.is_processing_pdf = False

        self.idx_pdf_entrada = tk.StringVar()
        self.idx_txt_indice = tk.StringVar()
        self.idx_pdf_salida = tk.StringVar()
        self.is_processing_idx = False

        self._build_top_menu()
        self._build_main_ui()

    def _build_top_menu(self):
        self.menubar = tk.Menu(self.root)
        self.menu_archivo = tk.Menu(self.menubar, tearoff=0)
        self.menu_archivo.add_command(label="Salir", command=self.salir_aplicacion, accelerator="Alt+F4")
        self.menubar.add_cascade(label="Archivo", menu=self.menu_archivo)
        self.root.config(menu=self.menubar)

    def salir_aplicacion(self):
        if messagebox.askokcancel("Salir", "¿Deseas cerrar la suite?"):
            self.root.destroy()

    def _build_main_ui(self):
        self.main_notebook = ttk.Notebook(self.root)
        self.main_notebook.pack(fill="both", expand=True, padx=10, pady=10)

        self.tab_csv_main = ttk.Frame(self.main_notebook)
        self.main_notebook.add(self.tab_csv_main, text=" ⚡ Motor de Datos (Big Data) ")

        self.tab_pdf_main = ttk.Frame(self.main_notebook)
        self.main_notebook.add(self.tab_pdf_main, text=" 🛠️ PDF Studio ")

        self.tab_riesgo_main = ttk.Frame(self.main_notebook)
        self.main_notebook.add(self.tab_riesgo_main, text=" 🛡️ Análisis de Riesgo y Muestreo ")

        self._build_csv_subnotebook()
        self._build_pdf_subnotebook()
        self._build_riesgo_demo_ui()

    def _build_csv_subnotebook(self):
        self.csv_notebook = ttk.Notebook(self.tab_csv_main)
        self.csv_notebook.pack(fill="both", expand=True, padx=5, pady=5)

        self.tab_separador = ttk.Frame(self.csv_notebook)
        self.tab_resumen = ttk.Frame(self.csv_notebook)
        self.tab_convertidor = ttk.Frame(self.csv_notebook)

        self.csv_notebook.add(self.tab_separador, text=" 📂 Separar por Cuenta ")
        self.csv_notebook.add(self.tab_resumen, text=" 📊 Resumen / Balanza ")
        self.csv_notebook.add(self.tab_convertidor, text=" ⚡ Convertidor ➔ Parquet ")

        self.setup_ui_separador()
        self.setup_ui_resumen()
        self.setup_ui_convertidor()
        
    def _obtener_lazy_frame(self, paths, col_dtypes=None):
        if not paths:
            return None

        # 1. Mapear los tipos de la interfaz a tipos nativos de Polars
        overrides_polars = {}
        if col_dtypes:
            for col_name, tipo_val in col_dtypes.items():
                # Si ya es un tipo de Polars se asigna directo, si es string de la UI se mapea
                if tipo_val in ("Texto", "str", "string"):
                    overrides_polars[col_name] = pl.Utf8
                elif tipo_val in ("Float (Decimal)", "float", "float64"):
                    overrides_polars[col_name] = pl.Float64
                elif tipo_val in ("Entero (Int)", "int", "int64"):
                    overrides_polars[col_name] = pl.Int64
                elif isinstance(tipo_val, pl.DataType):
                    overrides_polars[col_name] = tipo_val

        ext = os.path.splitext(paths[0])[1].lower()
        if ext == ".parquet":
            lf = pl.scan_parquet(paths)
            cols_originales = list(lf.collect_schema().keys())
            cols_limpias = {c: c.strip() for c in cols_originales}
            
            # Si es Parquet y se especificaron overrides, hacer cast de las columnas
            if overrides_polars:
                exprs = [
                    pl.col(c).cast(overrides_polars[c]) 
                    for c in overrides_polars if c in cols_originales
                ]
                if exprs:
                    lf = lf.with_columns(exprs)
                    
            return lf.rename(cols_limpias)
        else:
            if len(paths) == 1:
                clean_path, sep = obtener_archivo_limpio(paths[0])
                lf = pl.scan_csv(
                    clean_path, 
                    separator=sep, 
                    schema_overrides=overrides_polars if overrides_polars else None,
                    ignore_errors=True,
                    encoding="utf8"  # Se cambia 'utf8-lossy' por 'utf8' estricto
                )
                cols_originales = list(lf.collect_schema().keys())
                cols_limpias = {c: c.strip() for c in cols_originales}
                return lf.rename(cols_limpias)
            else:
                frames = []
                for p in paths:
                    clean_path, sep = obtener_archivo_limpio(p)
                    lf = pl.scan_csv(
                        clean_path, 
                        separator=sep, 
                        schema_overrides=overrides_polars if overrides_polars else None,
                        ignore_errors=True,
                        encoding="utf8"  # Se cambia 'utf8-lossy' por 'utf8' estricto
                    )
                    cols_originales = list(lf.collect_schema().keys())
                    cols_limpias = {c: c.strip() for c in cols_originales}
                    frames.append(lf.rename(cols_limpias))
                return pl.concat(frames)

    def _limpiar_nombres_columnas(self, raw_cols):
        columnas = []
        for idx, col in enumerate(raw_cols):
            nombre = str(col).strip() if col is not None else ""
            if not nombre or nombre.startswith("_duplicated"):
                nombre = f"Columna_{idx+1}"
            columnas.append(nombre)
        return columnas

    def _actualizar_ui_tipos_columnas(self, columnas):
        """
        Crea dinámicamente desplegables (Combobox) para asignar el tipo de dato
        a cada columna detectada en el archivo cargado.
        """
        # Inicializar o reiniciar el diccionario de tipos
        self.tipos_columnas = {}
    
        # Limpiar panel previo si existe
        if hasattr(self, 'frame_tipos') and self.frame_tipos.winfo_exists():
            self.frame_tipos.destroy()
    
        # Contenedor principal para la asignación de tipos estilo Power Query
        self.frame_tipos = ttk.LabelFrame(self.tab_separador, text=" Asignación de Tipos de Datos (Power Query) ", padding=10)
        self.frame_tipos.pack(fill="x", padx=15, pady=10)
    
        # Canvas con scrollbar por si hay muchas columnas
        canvas = tk.Canvas(self.frame_tipos, height=120)
        scrollbar = ttk.Scrollbar(self.frame_tipos, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
    
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
    
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
    
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
        # Crear fila para cada columna
        opciones_tipo = ["Texto", "Float (Decimal)", "Entero (Int)", "Fecha"]
    
        for idx, col in enumerate(columnas):
            lbl = ttk.Label(scrollable_frame, text=f"{col}:", font=("Segoe UI", 9, "bold"))
            lbl.grid(row=idx, column=0, padx=5, pady=3, sticky="w")
    
            combo = ttk.Combobox(scrollable_frame, values=opciones_tipo, state="readonly", width=18)
            
            # Detección predeterminada inteligente
            if "fecha" in col.lower() or "date" in col.lower():
                combo.set("Fecha")
            elif any(term in col.lower() for term in ["monto", "debe", "haber", "saldo", "amount"]):
                combo.set("Float (Decimal)")
            else:
                combo.set("Texto")
    
            combo.grid(row=idx, column=1, padx=5, pady=3, sticky="e")
    
            # Guardar la elección en el diccionario
            self.tipos_columnas[col] = combo.get()
            combo.bind("<<ComboboxSelected>>", lambda e, c=col, cb=combo: self._on_tipo_cambiado(c, cb.get()))
    
    def _on_tipo_cambiado(self, columna, nuevo_tipo):
        """Actualiza la selección en el diccionario interno."""
        self.tipos_columnas[columna] = nuevo_tipo

    def setup_ui_separador(self):
        self.file_display_sep = tk.StringVar()
        self.col_name_sep = tk.StringVar()
        self.out_dir_sep = tk.StringVar()
        self.format_sep = tk.StringVar(value="csv")

        tk.Label(self.tab_separador, text="1. Selecciona Archivo(s) de Entrada (CSV, Excel o Parquet):", font=("Arial", 10, "bold")).pack(pady=(15, 5))
        f_file = tk.Frame(self.tab_separador)
        f_file.pack()
        tk.Entry(f_file, textvariable=self.file_display_sep, width=50, state="readonly").pack(side=tk.LEFT, padx=5)
        tk.Button(f_file, text="Examinar...", command=self.seleccionar_archivos_sep).pack(side=tk.LEFT)

        tk.Label(self.tab_separador, text="2. Columna para Separar Cuentas:", font=("Arial", 10, "bold")).pack(pady=(15, 5))
        self.combo_col_sep = ttk.Combobox(self.tab_separador, textvariable=self.col_name_sep, state="readonly", width=47)
        self.combo_col_sep.pack()

        tk.Label(self.tab_separador, text="3. Formato de Salida de los Sub-archivos:", font=("Arial", 10, "bold")).pack(pady=(15, 5))
        f_fmt = tk.Frame(self.tab_separador)
        f_fmt.pack()
        tk.Radiobutton(f_fmt, text="CSV (.csv)", variable=self.format_sep, value="csv", font=("Arial", 9)).pack(side=tk.LEFT, padx=10)
        tk.Radiobutton(f_fmt, text="Excel (.xlsx)", variable=self.format_sep, value="xlsx", font=("Arial", 9)).pack(side=tk.LEFT, padx=10)
        tk.Radiobutton(f_fmt, text="Parquet (.parquet)", variable=self.format_sep, value="parquet", font=("Arial", 9)).pack(side=tk.LEFT, padx=10)

        tk.Label(self.tab_separador, text="4. Carpeta de Destino:", font=("Arial", 10, "bold")).pack(pady=(15, 5))
        f_dir = tk.Frame(self.tab_separador)
        f_dir.pack()
        tk.Entry(f_dir, textvariable=self.out_dir_sep, width=50, state="readonly").pack(side=tk.LEFT, padx=5)
        tk.Button(f_dir, text="Examinar...", command=self.seleccionar_carpeta_sep).pack(side=tk.LEFT)

        self.btn_process_sep = tk.Button(self.tab_separador, text="▶ DIVIDIR ARCHIVOS", command=self.iniciar_proceso_sep, bg="#4CAF50", fg="white", font=("Arial", 11, "bold"), pady=5, padx=20)
        self.btn_process_sep.pack(pady=20)

        self.status_sep = tk.Label(self.tab_separador, text="Esperando instrucciones...", fg="gray")
        self.status_sep.pack()

    def seleccionar_archivos_sep(self):
        paths = filedialog.askopenfilenames(filetypes=[("Archivos Soportados", "*.csv *.xlsx *.xlsm *.parquet"), ("Archivos CSV", "*.csv"), ("Archivos Excel", "*.xlsx *.xlsm"), ("Archivos Parquet", "*.parquet")])
        if paths:
            self.paths_sep = list(paths)
            self.file_display_sep.set(paths[0] if len(paths) == 1 else f"📁 {len(paths)} archivos seleccionados")
            self.out_dir_sep.set(os.path.join(os.path.dirname(paths[0]), "Movimientos_Separados"))
            
            try:
                # Se obtiene una muestra rápida de las primeras 5 filas con fetch(5)
                df_preview = self._obtener_lazy_frame(self.paths_sep).fetch(5)
                columnas_detectadas = df_preview.columns
                
                # 1. Actualizar el dropdown con las columnas detectadas
                self.combo_col_sep['values'] = columnas_detectadas
                if columnas_detectadas:
                    self.combo_col_sep.current(0)
                
                # 2. Generar/actualizar la lista visual de tipos de datos estilo Power Query
                self._actualizar_ui_tipos_columnas(columnas_detectadas)
                
                self.status_sep.config(text=f"Columnas detectadas correctamente en {len(paths)} archivo(s).", fg="green")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo leer el archivo:\n{e}")

    def seleccionar_carpeta_sep(self):
        path = filedialog.askdirectory()
        if path:
            self.out_dir_sep.set(path)

    def iniciar_proceso_sep(self):
        if not self.paths_sep or not self.col_name_sep.get() or not self.out_dir_sep.get():
            messagebox.showwarning("Datos Incompletos", "Completa todos los campos.")
            return

        self.btn_process_sep.config(state=tk.DISABLED)
        self.status_sep.config(text="Procesando datos con Polars...", fg="blue")
        threading.Thread(target=self._separar_datos_thread, daemon=True).start()


    def _separar_datos_thread(self):
        col = self.col_name_sep.get()
        carpeta_out = self.out_dir_sep.get()
        formato = self.format_sep.get()
        try:
            os.makedirs(carpeta_out, exist_ok=True)
            lazy_base = self._obtener_lazy_frame(self.paths_sep, {col: pl.String})
    
            # --- PIPELINE DE LIMPIEZA DINÁMICO Y RELLENO CON POLARS ---
            lazy_procesado = (
                lazy_base
                .with_columns(
                    # 1. Limpiar espacios y normalizar celdas vacías
                    pl.col(col)
                    .cast(pl.Utf8)
                    .str.strip_chars()
                    .replace("", None)
                )
                # 2. Descartar filas con cualquier variación de la palabra 'total'
                .filter(
                    ~pl.col(col).str.contains(r"(?i)\btotal\b").fill_null(False)
                )
                # 3. Propagar la cuenta hacia abajo (Forward Fill)
                .with_columns(
                    pl.col(col).forward_fill()
                )
                # 4. Eliminar filas no asignadas
                .filter(pl.col(col).is_not_null())
            )
    
            # Obtener el catálogo único de cuentas resultantes
            df_cuentas = lazy_procesado.select(pl.col(col)).unique().collect()
            cuentas = [str(c) for c in df_cuentas[col].to_list() if c is not None]
            total = len(cuentas)
    
            LIMITE_EXCEL = 1_000_000  # Margen de seguridad bajo el límite de 1,048,576 filas
    
            for i, cuenta in enumerate(cuentas):
                self.root.after(0, self.status_sep.config, {"text": f"Procesando ({i+1}/{total}): {cuenta}"})
                safe_name = "".join(c for c in cuenta if c.isalnum() or c in (' ', '_', '-')).strip()
                
                # Recolectar datos de la cuenta y aplicar la limpieza de formatos
                df_grupo = lazy_procesado.filter(pl.col(col) == cuenta).collect()
                if hasattr(self, 'tipos_columnas') and self.tipos_columnas:
                    df_limpio = aplicar_formato_personalizado(df_grupo, self.tipos_columnas)
                else:
                    df_limpio = aplicar_formato_limpio(df_grupo)

                # --- BLOQUE DE EXPORTACIÓN CON PROTECCIÓN PARA EXCEL ---
                if formato == "csv":
                    # Se utiliza utf-8-sig para preservar caracteres especiales en Excel
                    df_limpio.write_csv(os.path.join(carpeta_out, f"{safe_name}.csv"))
    
                elif formato == "parquet":
                    df_limpio.write_parquet(os.path.join(carpeta_out, f"{safe_name}.parquet"))
    
                else:  # formato == "xlsx"
                    excel_path = os.path.join(carpeta_out, f"{safe_name}.xlsx")
                    total_filas = df_limpio.height
    
                    # Caso A: Si cabe en una sola hoja
                    if total_filas <= LIMITE_EXCEL:
                        df_limpio.write_excel(excel_path)
    
                    # Caso B: Si supera el límite, fragmenta en múltiples pestañas
                    else:
                        num_partes = math.ceil(total_filas / LIMITE_EXCEL)
                        
                        with xlsxwriter.Workbook(excel_path, {'constant_memory': True}) as workbook:
                            for parte in range(num_partes):
                                offset = parte * LIMITE_EXCEL
                                df_chunk = df_limpio.slice(offset, LIMITE_EXCEL)
                                
                                sheet_name = f"Parte_{parte + 1}"
                                worksheet = workbook.add_worksheet(sheet_name)
                                
                                # Escribir Encabezados
                                for col_num, col_name in enumerate(df_chunk.columns):
                                    worksheet.write(0, col_num, col_name)
                                
                                # Escribir Filas
                                for row_idx, row_data in enumerate(df_chunk.iter_rows()):
                                    for col_idx, val in enumerate(row_data):
                                        worksheet.write(row_idx + 1, col_idx, val)
    
            self.root.after(0, self.status_sep.config, {"text": "✅ Proceso finalizado.", "fg": "green"})
            self.root.after(0, messagebox.showinfo, "Éxito", f"Archivos exportados en:\n{carpeta_out}")
        except Exception as e:
            self.root.after(0, messagebox.showerror, "Error", f"Ocurrió un error:\n{e}")
            self.root.after(0, self.status_sep.config, {"text": "❌ Error.", "fg": "red"})
        finally:
            self.root.after(0, self.btn_process_sep.config, {"state": tk.NORMAL})

    def setup_ui_resumen(self):
        self.file_display_res = tk.StringVar()
        self.col_group_res = tk.StringVar()
        self.out_file_res = tk.StringVar()
        self.format_res = tk.StringVar(value="xlsx")
        # Variable para controlar si se excluyen las filas de 'TOTAL'
        self.limpiar_totales_res = tk.BooleanVar(value=True)

        tk.Label(self.tab_resumen, text="1. Selecciona Archivo(s) de Entrada (CSV, Excel o Parquet):", font=("Arial", 10, "bold")).pack(pady=(10, 2))
        f_file = tk.Frame(self.tab_resumen)
        f_file.pack()
        tk.Entry(f_file, textvariable=self.file_display_res, width=50, state="readonly").pack(side=tk.LEFT, padx=5)
        tk.Button(f_file, text="Examinar...", command=self.seleccionar_archivos_res).pack(side=tk.LEFT)

        tk.Label(self.tab_resumen, text="2. Columna para Agrupar (Ej: Cuenta Contable):", font=("Arial", 10, "bold")).pack(pady=(10, 2))
        self.combo_col_res = ttk.Combobox(self.tab_resumen, textvariable=self.col_group_res, state="readonly", width=47)
        self.combo_col_res.pack()

        tk.Label(self.tab_resumen, text="3. Columnas de Valor a Sumar (Mantén Ctrl para elegir varias):", font=("Arial", 10, "bold")).pack(pady=(10, 2))
        f_list = tk.Frame(self.tab_resumen)
        f_list.pack()
        self.listbox_vals = tk.Listbox(f_list, selectmode=tk.MULTIPLE, width=47, height=5, exportselection=0)
        scrollbar = tk.Scrollbar(f_list, orient="vertical", command=self.listbox_vals.yview)
        self.listbox_vals.config(yscrollcommand=scrollbar.set)
        self.listbox_vals.pack(side=tk.LEFT)
        scrollbar.pack(side=tk.RIGHT, fill="y")

        tk.Label(self.tab_resumen, text="4. Formato de Salida del Resumen:", font=("Arial", 10, "bold")).pack(pady=(10, 2))
        f_fmt = tk.Frame(self.tab_resumen)
        f_fmt.pack()
        tk.Radiobutton(f_fmt, text="Excel (.xlsx)", variable=self.format_res, value="xlsx", font=("Arial", 9)).pack(side=tk.LEFT, padx=15)
        tk.Radiobutton(f_fmt, text="CSV (.csv)", variable=self.format_res, value="csv", font=("Arial", 9)).pack(side=tk.LEFT, padx=15)

        # Casilla de verificación para filtrar filas de totales
        f_chk = tk.Frame(self.tab_resumen)
        f_chk.pack(pady=(5, 5))
        tk.Checkbutton(
            f_chk, 
            text="Excluir filas que contengan la palabra 'TOTAL' (Útil para balances contables)", 
            variable=self.limpiar_totales_res, 
            font=("Arial", 9)
        ).pack()

        tk.Label(self.tab_resumen, text="5. Guardar Resumen Como:", font=("Arial", 10, "bold")).pack(pady=(10, 2))
        f_out = tk.Frame(self.tab_resumen)
        f_out.pack()
        tk.Entry(f_out, textvariable=self.out_file_res, width=50, state="readonly").pack(side=tk.LEFT, padx=5)
        tk.Button(f_out, text="Examinar...", command=self.seleccionar_salida_res).pack(side=tk.LEFT)

        self.btn_process_res = tk.Button(self.tab_resumen, text="⚡ GENERAR BALANZA / RESUMEN", command=self.iniciar_proceso_res, bg="#2196F3", fg="white", font=("Arial", 11, "bold"), pady=5, padx=15)
        self.btn_process_res.pack(pady=15)

        self.status_res = tk.Label(self.tab_resumen, text="Esperando instrucciones...", fg="gray")
        self.status_res.pack()

    def seleccionar_archivos_res(self):
        paths = filedialog.askopenfilenames(filetypes=[("Archivos Soportados", "*.csv *.xlsx *.xlsm *.parquet"), ("Archivos CSV", "*.csv"), ("Archivos Excel", "*.xlsx *.xlsm"), ("Archivos Parquet", "*.parquet")])
        if paths:
            self.paths_res = list(paths)
            self.file_display_res.set(paths[0] if len(paths) == 1 else f"📁 {len(paths)} archivos seleccionados")
            self.out_file_res.set(os.path.join(os.path.dirname(paths[0]), "Resumen_Balanza_Contable.xlsx"))

            try:
                lazy_df = self._obtener_lazy_frame(self.paths_res)
                raw_cols = lazy_df.collect_schema().names()
                columnas = self._limpiar_nombres_columnas(raw_cols)

                self.combo_col_res['values'] = columnas
                if columnas:
                    self.combo_col_res.current(0)

                self.listbox_vals.delete(0, tk.END)
                for c in columnas:
                    self.listbox_vals.insert(tk.END, c)

                self.status_res.config(text=f"Columnas detectadas correctamente en {len(paths)} archivo(s).", fg="green")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo leer el archivo:\n{e}")

    def seleccionar_salida_res(self):
        fmt = self.format_res.get()
        ext = "*.xlsx" if fmt == "xlsx" else "*.csv"
        path = filedialog.asksaveasfilename(defaultextension=f".{fmt}", filetypes=[("Archivo " + fmt.upper(), ext)])
        if path:
            self.out_file_res.set(path)

    def iniciar_proceso_res(self):
        indices_vals = self.listbox_vals.curselection()
        if not self.paths_res or not self.col_group_res.get() or not self.out_file_res.get() or not indices_vals:
            messagebox.showwarning("Datos Incompletos", "Completa todos los campos.")
            return

        self.btn_process_res.config(state=tk.DISABLED)
        self.status_res.config(text="Generando resumen con Polars...", fg="blue")
        threading.Thread(target=self._generar_resumen_thread, daemon=True).start()
    
    def _generar_resumen_thread(self):
        col_agrupar = self.col_group_res.get()
        indices_vals = self.listbox_vals.curselection()
        cols_sumar = [self.listbox_vals.get(i) for i in indices_vals]
        archivo_out = self.out_file_res.get()
        formato = self.format_res.get()

        try:
            # 1. Obtener la base de datos
            lazy_base = self._obtener_lazy_frame(self.paths_res, {col_agrupar: pl.String})

            # 2. Colectar temporalmente para sanear las columnas con el algoritmo agnóstico
            df_base = lazy_base.collect()
            if hasattr(self, 'tipos_columnas') and self.tipos_columnas:
                df_saneado = aplicar_formato_personalizado(df_base, self.tipos_columnas)
            else:
                df_saneado = aplicar_formato_limpio(df_base)

            # 3. Expresiones de suma segura con casteo a Float64 y limpieza de comas
            exprs = [
                pl.col(c)
                .cast(pl.Utf8)                          # Asegura que se lea como texto para limpiar
                .str.replace_all(",", "")               # Elimina comas de formato de miles (ej. 1,500.00)
                .cast(pl.Float64, strict=False)         # Convierte a numérico sin colapsar por valores nulos
                .fill_null(0.0)                         # Trata textos no numéricos o vacíos como 0.0
                .sum()
                .round(2)
                .alias(c) 
                for c in cols_sumar
            ]

            # 4. Agrupar y procesar el resumen con control condicional del Checkbutton
            df_temp = df_saneado.with_columns(
                pl.col(col_agrupar).cast(pl.Utf8).str.strip_chars().replace("", None)
            )

            # Si el Checkbutton está activado, elimina las filas con la palabra 'total'
            if self.limpiar_totales_res.get():
                df_temp = df_temp.filter(~pl.col(col_agrupar).str.contains(r"(?i)\btotal\b").fill_null(False))

            df_resumen = (
                df_temp
                .with_columns(pl.col(col_agrupar).forward_fill())
                .filter(pl.col(col_agrupar).is_not_null())
                .group_by(col_agrupar)
                .agg(exprs)
                .sort(col_agrupar)
            )

            # 5. Exportar según formato con soporte de BOM para UTF-8 en CSV
            if formato == "csv":
                if not archivo_out.endswith(".csv"):
                    archivo_out = os.path.splitext(archivo_out)[0] + ".csv"
                
                # Escribir con BOM (\xef\xbb\xbf) para mantener símbolos intactos en Excel
                with open(archivo_out, "wb") as f:
                    f.write(b'\xef\xbb\xbf')
                    df_resumen.write_csv(f)
            else:
                if not archivo_out.endswith(".xlsx"):
                    archivo_out = os.path.splitext(archivo_out)[0] + ".xlsx"
                df_resumen.write_excel(archivo_out)

            self.root.after(0, self.status_res.config, {"text": "✅ Resumen generado exitosamente.", "fg": "green"})
            self.root.after(0, messagebox.showinfo, "Éxito", f"Procesadas {len(df_resumen)} cuentas únicas.\nGuardado en:\n{archivo_out}")
        except Exception as e:
            self.root.after(0, messagebox.showerror, "Error Fatal", f"Ocurrió un error:\n{e}")
            self.root.after(0, self.status_res.config, {"text": "❌ Error.", "fg": "red"})
        finally:
            self.root.after(0, self.btn_process_res.config, {"state": tk.NORMAL})

    def setup_ui_convertidor(self):
        self.file_display_conv = tk.StringVar()
        self.out_file_conv = tk.StringVar()

        tk.Label(self.tab_convertidor, text="1. Selecciona el o los Archivos a Convertir (CSV o Excel):", font=("Arial", 10, "bold")).pack(pady=(20, 5))
        f_file = tk.Frame(self.tab_convertidor)
        f_file.pack()
        tk.Entry(f_file, textvariable=self.file_display_conv, width=50, state="readonly").pack(side=tk.LEFT, padx=5)
        tk.Button(f_file, text="Examinar...", command=self.seleccionar_archivos_conv).pack(side=tk.LEFT)

        tk.Label(self.tab_convertidor, text="2. Archivo Parquet de Destino (Consolidado):", font=("Arial", 10, "bold")).pack(pady=(20, 5))
        f_out = tk.Frame(self.tab_convertidor)
        f_out.pack()
        tk.Entry(f_out, textvariable=self.out_file_conv, width=50, state="readonly").pack(side=tk.LEFT, padx=5)
        tk.Button(f_out, text="Guardar como...", command=self.seleccionar_salida_conv).pack(side=tk.LEFT)

        self.btn_process_conv = tk.Button(self.tab_convertidor, text="⚡ CONVERTIR Y CONSOLIDAR A PARQUET", command=self.iniciar_proceso_conv, bg="#9C27B0", fg="white", font=("Arial", 11, "bold"), pady=5, padx=15)
        self.btn_process_conv.pack(pady=30)

        self.status_conv = tk.Label(self.tab_convertidor, text="Esperando instrucciones...", fg="gray")
        self.status_conv.pack()

    def seleccionar_archivos_conv(self):
        paths = filedialog.askopenfilenames(filetypes=[("Archivos Soportados", "*.csv *.xlsx *.xlsm"), ("Archivos CSV", "*.csv"), ("Archivos Excel", "*.xlsx *.xlsm")])
        if paths:
            self.paths_conv = list(paths)
            self.file_display_conv.set(paths[0] if len(paths) == 1 else f"📁 {len(paths)} archivos seleccionados")
            self.out_file_conv.set(os.path.join(os.path.dirname(paths[0]), "Datos_Contables_Consolidados.parquet"))
            self.status_conv.config(text=f"{len(paths)} archivo(s) listos para conversión.", fg="green")

    def seleccionar_salida_conv(self):
        path = filedialog.asksaveasfilename(defaultextension=".parquet", filetypes=[("Archivo Parquet", "*.parquet")])
        if path:
            self.out_file_conv.set(path)

    def iniciar_proceso_conv(self):
        if not self.paths_conv or not self.out_file_conv.get():
            messagebox.showwarning("Datos Incompletos", "Selecciona los archivos de origen (CSV o Excel) y el destino .parquet.")
            return

        self.btn_process_conv.config(state=tk.DISABLED)
        self.status_conv.config(text="Transformando y comprimiendo a Parquet...", fg="blue")
        threading.Thread(target=self._convertir_parquet_thread, daemon=True).start()

    def _convertir_parquet_thread(self):
        archivo_out = self.out_file_conv.get()
        try:
            lazy_df = self._obtener_lazy_frame(self.paths_conv)
            lazy_df.sink_parquet(archivo_out, compression="zstd")

            self.root.after(0, self.status_conv.config, {"text": "✅ Conversión a Parquet completada.", "fg": "green"})
            self.root.after(0, messagebox.showinfo, "Éxito", f"Se han consolidado {len(self.paths_conv)} archivo(s) en el archivo Parquet:\n{archivo_out}")
        except Exception as e:
            self.root.after(0, messagebox.showerror, "Error Fatal", f"Ocurrió un error:\n{e}")
            self.root.after(0, self.status_conv.config, {"text": "❌ Error.", "fg": "red"})
        finally:
            self.root.after(0, self.btn_process_conv.config, {"state": tk.NORMAL})

    # ==========================================
    # PESTAÑA PRINCIPAL: TRATAMIENTO DE PDFS
    # ==========================================
    def _build_pdf_subnotebook(self):
        self.pdf_notebook = ttk.Notebook(self.tab_pdf_main)
        self.pdf_notebook.pack(fill="both", expand=True, padx=5, pady=5)

        self.tab_pdf_divisor = ttk.Frame(self.pdf_notebook)
        self.tab_pdf_indice = ttk.Frame(self.pdf_notebook)

        self.pdf_notebook.add(self.tab_pdf_divisor, text=" 📂 Divisor por Tamaño / Límite ")
        self.pdf_notebook.add(self.tab_pdf_indice, text=" 🔖 Índices y Marcadores ")

        self.setup_ui_pdf_divisor()
        self.setup_ui_pdf_indice()

    def setup_ui_pdf_divisor(self):
        file_frame = ttk.LabelFrame(self.tab_pdf_divisor, text=" Archivo PDF de Entrada ", padding=10)
        file_frame.pack(fill="x", padx=15, pady=5)

        ttk.Entry(file_frame, textvariable=self.pdf_path).pack(side="left", fill="x", expand=True, padx=(0, 5))
        ttk.Button(file_frame, text="Buscar PDF...", command=self._browse_pdf).pack(side="right")

        out_frame = ttk.LabelFrame(self.tab_pdf_divisor, text=" Carpeta de Salida ", padding=10)
        out_frame.pack(fill="x", padx=15, pady=5)

        ttk.Entry(out_frame, textvariable=self.output_dir_pdf).pack(side="left", fill="x", expand=True, padx=(0, 5))
        ttk.Button(out_frame, text="Buscar Carpeta...", command=self._browse_output_dir_pdf).pack(side="right")

        config_frame = ttk.LabelFrame(self.tab_pdf_divisor, text=" Configuración de Límite ", padding=10)
        config_frame.pack(fill="x", padx=15, pady=5)

        ttk.Label(config_frame, text="Tamaño máximo por parte (MB):").pack(side="left", padx=(0, 10))
        ttk.Spinbox(config_frame, from_=0.5, to=500.0, increment=0.5, textvariable=self.max_size_mb, width=10).pack(side="left")

        save_frame = ttk.LabelFrame(self.tab_pdf_divisor, text=" Opción de Guardado ", padding=10)
        save_frame.pack(fill="x", padx=15, pady=5)

        ttk.Radiobutton(save_frame, text="Solo partes PDF", variable=self.export_mode_pdf, value="pdf").pack(side="left", padx=(0, 15))
        ttk.Radiobutton(save_frame, text="Solo comprimido (.zip)", variable=self.export_mode_pdf, value="zip").pack(side="left", padx=(0, 15))
        ttk.Radiobutton(save_frame, text="Ambos (.pdf y .zip)", variable=self.export_mode_pdf, value="both").pack(side="left")

        self.btn_process_pdf = ttk.Button(self.tab_pdf_divisor, text="▶ DIVIDIR Y PROCESAR PDF", command=self._start_pdf_process_thread)
        self.btn_process_pdf.pack(fill="x", padx=15, pady=10)

        log_frame = ttk.LabelFrame(self.tab_pdf_divisor, text=" Consola de Monitorización y Detalles del PDF ", padding=10)
        log_frame.pack(fill="both", expand=True, padx=15, pady=(5, 15))

        self.log_widget_pdf = scrolledtext.ScrolledText(log_frame, wrap="word", state="disabled", height=8, bg="#1e1e1e", fg="#d4d4d4")
        self.log_widget_pdf.pack(fill="both", expand=True)

    def _browse_pdf(self):
        file_selected = filedialog.askopenfilename(title="Seleccionar PDF", filetypes=[("Archivos PDF", "*.pdf")])
        if file_selected:
            self.pdf_path.set(file_selected)
            if not self.output_dir_pdf.get():
                self.output_dir_pdf.set(os.path.dirname(file_selected))

    def _browse_output_dir_pdf(self):
        dir_selected = filedialog.askdirectory(title="Seleccionar Carpeta de Salida")
        if dir_selected:
            self.output_dir_pdf.set(dir_selected)

    def log_pdf(self, text):
        self.log_widget_pdf.config(state="normal")
        self.log_widget_pdf.insert(tk.END, text + "\n")
        self.log_widget_pdf.see(tk.END)
        self.log_widget_pdf.config(state="disabled")

    def _clear_log_pdf(self):
        self.log_widget_pdf.config(state="normal")
        self.log_widget_pdf.delete("1.0", tk.END)
        self.log_widget_pdf.config(state="disabled")

    def _start_pdf_process_thread(self):
        if self.is_processing_pdf:
            return

        pdf_path = self.pdf_path.get().strip()
        output_dir = self.output_dir_pdf.get().strip()
        max_mb = self.max_size_mb.get()

        if not pdf_path or not os.path.exists(pdf_path):
            messagebox.showerror("Error", "Por favor selecciona un archivo PDF válido.")
            return

        if not output_dir or not os.path.exists(output_dir):
            messagebox.showerror("Error", "Por favor selecciona una carpeta de salida válida.")
            return

        if max_mb <= 0:
            messagebox.showerror("Error", "El tamaño máximo debe ser mayor que 0 MB.")
            return

        self._clear_log_pdf()
        self.is_processing_pdf = True
        self.btn_process_pdf.config(state="disabled")

        thread = threading.Thread(target=self._run_pdf_process, args=(pdf_path, output_dir, max_mb), daemon=True)
        thread.start()

    def get_pdf_size(self, pdf_path):
        return os.path.getsize(pdf_path) / (1024 * 1024)

    def compress_pdf_lossless(self, input_path, output_path):
        original_size = self.get_pdf_size(input_path)
        if os.path.abspath(input_path) == os.path.abspath(output_path):
            self.log_pdf("⚠️ Archivo de entrada y salida son iguales, omitiendo compresión.")
            return original_size
        try:
            with pikepdf.open(input_path) as pdf:
                pdf.save(output_path, compress_streams=True, object_stream_mode=pikepdf.ObjectStreamMode.preserve)
            new_size = self.get_pdf_size(output_path)
            reduction = ((original_size - new_size) / original_size) * 100
            self.log_pdf(f"📊 Tamaño original: {original_size:.2f} MB")
            self.log_pdf(f"📊 Tamaño tras compresión: {new_size:.2f} MB")
            self.log_pdf(f"📊 Reducción: {reduction:.1f}%")
            return new_size
        except Exception as e:
            self.log_pdf(f"❌ Error en compresión: {e}")
            if os.path.abspath(input_path) != os.path.abspath(output_path):
                shutil.copy2(input_path, output_path)
            return original_size

    def calculate_parts_needed(self, compressed_size_mb, max_size_mb):
        if compressed_size_mb <= max_size_mb:
            return 1
        return math.ceil(compressed_size_mb / max_size_mb)

    def find_optimal_cut_point(self, pdf_path, start_page, total_pages, max_size_mb):
        temp_dir = os.path.join(self.output_dir_pdf.get(), "temp_opt")
        os.makedirs(temp_dir, exist_ok=True)
        reader = PdfReader(pdf_path)
        best_end_page = start_page
        best_size = 0

        self.log_pdf(f"   🔍 Buscando combinación óptima desde página {start_page + 1}...")

        for test_end_page in range(start_page + 1, total_pages + 1):
            writer = PdfWriter()
            writer.append(reader, pages=(start_page, test_end_page))

            temp_path = os.path.join(temp_dir, "test_part.pdf")
            with open(temp_path, 'wb') as f:
                writer.write(f)

            part_size = self.get_pdf_size(temp_path)

            if part_size <= max_size_mb:
                best_end_page = test_end_page
                best_size = part_size
                if part_size >= max_size_mb * 0.95:
                    self.log_pdf(f"   ✅ Corte óptimo: págs {start_page+1}-{test_end_page} ({part_size:.2f} MB)")
                    break
            else:
                if best_end_page > start_page:
                    self.log_pdf(f"   ✅ Mejor combinación: págs {start_page+1}-{best_end_page} ({best_size:.2f} MB)")
                else:
                    best_end_page = start_page + 1
                    writer_forced = PdfWriter()
                    writer_forced.append(reader, pages=(start_page, start_page + 1))
                    temp_forced_path = os.path.join(temp_dir, "forced_test.pdf")
                    with open(temp_forced_path, 'wb') as f:
                        writer_forced.write(f)
                    forced_size = self.get_pdf_size(temp_forced_path)
                    self.log_pdf(f"   ⚠️ Forzando corte mínimo (1 pág): pág {start_page+1} ({forced_size:.2f} MB)")
                break
        else:
            if best_end_page > start_page:
                self.log_pdf(f"   ✅ Última combinación: págs {start_page+1}-{best_end_page} ({best_size:.2f} MB)")

        shutil.rmtree(temp_dir, ignore_errors=True)
        return best_end_page, best_size

    def split_pdf_by_max_size(self, pdf_path, output_dir, num_parts, max_size_mb):
        if not os.path.exists(pdf_path):
            self.log_pdf(f"❌ Error: No se encuentra {pdf_path}")
            return [], 0

        reader = PdfReader(pdf_path)
        total_pages = len(reader.pages)
        self.log_pdf(f"📊 Total de páginas del documento: {total_pages}")

        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
        if base_name.endswith('_comprimido'):
            base_name = base_name.replace('_comprimido', '')

        output_files = []
        current_page = 0
        actual_parts = 0

        while current_page < total_pages:
            actual_parts += 1
            self.log_pdf(f"\n📦 Procesando parte {actual_parts}...")

            end_page, part_size = self.find_optimal_cut_point(
                pdf_path, current_page, total_pages, max_size_mb
            )

            writer = PdfWriter()
            writer.append(reader, pages=(current_page, end_page))

            output_filename = f"{base_name}_parte{actual_parts}.pdf"
            output_path = os.path.join(output_dir, output_filename)
            with open(output_path, 'wb') as output_file:
                writer.write(output_file)

            final_size = self.get_pdf_size(output_path)
            status = "✅ DENTRO DEL LÍMITE" if final_size <= max_size_mb else "⚠️ EXCEDE LÍMITE"
            self.log_pdf(f"   {status} Parte {actual_parts}: págs {current_page+1}-{end_page}, tamaño: {final_size:.2f} MB")

            output_files.append(output_path)
            current_page = end_page

            if current_page >= total_pages:
                break

        return output_files, actual_parts

    def _run_pdf_process(self, pdf_path, output_dir, max_size_mb):
        try:
            export_mode = self.export_mode_pdf.get()
            self.log_pdf("🚀 Iniciando procesamiento del PDF...")
            original_size = self.get_pdf_size(pdf_path)
            self.log_pdf(f"📄 Archivo: {os.path.basename(pdf_path)} ({original_size:.2f} MB)")

            temp_dir = os.path.join(output_dir, "pdf_temp_process")
            os.makedirs(temp_dir, exist_ok=True)
            base_name = os.path.splitext(os.path.basename(pdf_path))[0]

            self.log_pdf("\n🔄 Aplicando compresión sin pérdida (PikePDF)...")
            compressed_path = os.path.join(temp_dir, f"{base_name}_comprimido.pdf")
            final_size = self.compress_pdf_lossless(pdf_path, compressed_path)

            parts_needed_estimate = self.calculate_parts_needed(final_size, max_size_mb)

            if parts_needed_estimate == 1 and final_size <= max_size_mb:
                self.log_pdf(f"\n✅ {final_size:.2f} MB ≤ {max_size_mb} MB → El archivo ya cumple con el límite establecido.")
                final_filename = f"{base_name}_comprimido.pdf"
                final_path = os.path.join(output_dir, final_filename)
                shutil.copy2(compressed_path, final_path)

                if export_mode == "zip":
                    zip_filename = os.path.join(output_dir, f"{base_name}_comprimido.zip")
                    with zipfile.ZipFile(zip_filename, 'w') as zipf:
                        zipf.write(final_path, os.path.basename(final_path))
                    os.remove(final_path)
                    self.log_pdf(f"📦 Guardado exclusivamente en archivo ZIP: {zip_filename}")
                elif export_mode == "both":
                    zip_filename = os.path.join(output_dir, f"{base_name}_comprimido.zip")
                    with zipfile.ZipFile(zip_filename, 'w') as zipf:
                        zipf.write(final_path, os.path.basename(final_path))
                    self.log_pdf(f"🎉 Archivos generados correctamente en PDF y ZIP: {final_path}")
                else:
                    self.log_pdf(f"🎉 Guardado únicamente en formato PDF: {final_path}")
            else:
                self.log_pdf(f"\n⚠️ {final_size:.2f} MB > {max_size_mb} MB → Iniciando división óptima...")
                parts, actual_parts = self.split_pdf_by_max_size(compressed_path, output_dir, parts_needed_estimate, max_size_mb)

                if export_mode in ("zip", "both"):
                    zip_filename = os.path.join(output_dir, f"{base_name}_partes.zip")
                    with zipfile.ZipFile(zip_filename, 'w') as zipf:
                        for part in parts:
                            zipf.write(part, os.path.basename(part))
                    self.log_pdf(f"\n📦 Archivo ZIP empaquetado: {zip_filename}")

                if export_mode == "zip":
                    for part in parts:
                        if os.path.exists(part):
                            os.remove(part)
                    self.log_pdf("🗑️ Partes PDF individuales eliminadas (sólo se conserva el ZIP final).")
                elif export_mode == "pdf":
                    self.log_pdf("\n📄 Se conservan únicamente las partes PDF resultantes.")

            shutil.rmtree(temp_dir, ignore_errors=True)
            self.log_pdf("\n✅ ¡Proceso de PDF completado exitosamente!")
            self.root.after(0, messagebox.showinfo, "Éxito", "El PDF ha sido procesado y dividido correctamente.")

        except Exception as e:
            self.log_pdf(f"\n❌ Error durante el procesamiento: {str(e)}")
            self.root.after(0, messagebox.showerror, "Error", f"Ocurrió un error inesperado:\n{str(e)}")
        finally:
            self.is_processing_pdf = False
            self.root.after(0, lambda: self.btn_process_pdf.config(state="normal"))

    def setup_ui_pdf_indice(self):
        frame_inputs = ttk.LabelFrame(self.tab_pdf_indice, text=" Configuración de Archivos ", padding=10)
        frame_inputs.pack(fill="x", padx=15, pady=5)
        frame_inputs.columnconfigure(1, weight=1)

        ttk.Label(frame_inputs, text="PDF de Entrada:").grid(row=0, column=0, sticky="w", pady=4)
        ttk.Entry(frame_inputs, textvariable=self.idx_pdf_entrada).grid(row=0, column=1, sticky="ew", padx=5, pady=4)
        ttk.Button(frame_inputs, text="Buscar...", command=self._buscar_idx_pdf).grid(row=0, column=2, padx=2, pady=4)

        ttk.Label(frame_inputs, text="Archivo TXT (Índice):").grid(row=1, column=0, sticky="w", pady=4)
        ttk.Entry(frame_inputs, textvariable=self.idx_txt_indice).grid(row=1, column=1, sticky="ew", padx=5, pady=4)
        ttk.Button(frame_inputs, text="Buscar...", command=self._buscar_idx_txt).grid(row=1, column=2, padx=2, pady=4)

        ttk.Label(frame_inputs, text="PDF de Salida:").grid(row=2, column=0, sticky="w", pady=4)
        ttk.Entry(frame_inputs, textvariable=self.idx_pdf_salida).grid(row=2, column=1, sticky="ew", padx=5, pady=4)
        ttk.Button(frame_inputs, text="Guardar en...", command=self._guardar_idx_pdf).grid(row=2, column=2, padx=2, pady=4)

        self.btn_process_idx = ttk.Button(
            self.tab_pdf_indice, 
            text="⚡ GENERAR ÍNDICE, VÍNCULOS Y MARCADORES", 
            command=self._start_idx_process_thread
        )
        self.btn_process_idx.pack(fill="x", padx=15, pady=8)

        log_frame = ttk.LabelFrame(self.tab_pdf_indice, text=" Consola de Monitorización y Diagnóstico ", padding=10)
        log_frame.pack(fill="both", expand=True, padx=15, pady=(5, 15))

        self.log_widget_idx = scrolledtext.ScrolledText(
            log_frame, 
            wrap="word", 
            state="disabled", 
            height=10, 
            bg="#1e1e1e", 
            fg="#d4d4d4"
        )
        self.log_widget_idx.pack(fill="both", expand=True)

    def _buscar_idx_pdf(self):
        ruta = filedialog.askopenfilename(filetypes=[("Archivos PDF", "*.pdf")])
        if ruta: 
            self.idx_pdf_entrada.set(ruta)
            if not self.idx_pdf_salida.get():
                base, ext = os.path.splitext(ruta)
                self.idx_pdf_salida.set(f"{base}_con_indice{ext}")

    def _buscar_idx_txt(self):
        ruta = filedialog.askopenfilename(filetypes=[("Archivos de texto", "*.txt")])
        if ruta: 
            self.idx_txt_indice.set(ruta)

    def _guardar_idx_pdf(self):
        ruta = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("Archivos PDF", "*.pdf")])
        if ruta: 
            self.idx_pdf_salida.set(ruta)

    def log_idx(self, text):
        self.log_widget_idx.config(state="normal")
        self.log_widget_idx.insert(tk.END, text + "\n")
        self.log_widget_idx.see(tk.END)
        self.log_widget_idx.config(state="disabled")

    def _clear_log_idx(self):
        self.log_widget_idx.config(state="normal")
        self.log_widget_idx.delete("1.0", tk.END)
        self.log_widget_idx.config(state="disabled")

    def _start_idx_process_thread(self):
        if self.is_processing_idx:
            return

        pdf_in = self.idx_pdf_entrada.get().strip()
        txt_in = self.idx_txt_indice.get().strip()
        pdf_out = self.idx_pdf_salida.get().strip()

        if not pdf_in or not os.path.exists(pdf_in):
            messagebox.showerror("Error", "Por favor selecciona un archivo PDF de entrada válido.")
            return

        if not txt_in or not os.path.exists(txt_in):
            messagebox.showerror("Error", "Por favor selecciona un archivo TXT de índice válido.")
            return

        if not pdf_out:
            messagebox.showerror("Error", "Por favor especifica la ruta del PDF de salida.")
            return

        self._clear_log_idx()
        self.is_processing_idx = True
        self.btn_process_idx.config(state="disabled")

        thread = threading.Thread(
            target=self._run_idx_process, 
            args=(pdf_in, txt_in, pdf_out), 
            daemon=True
        )
        thread.start()

    def _run_idx_process(self, pdf_entrada, txt_indice, pdf_salida):
        tiempo_inicio = time.perf_counter()
        try:
            self.log_idx("==================================================")
            self.log_idx("🚀 INICIANDO GENERACIÓN DE ÍNDICE Y MARCADORES")
            self.log_idx("==================================================")

            self.log_idx(f"📖 Leyendo estructura desde: {os.path.basename(txt_indice)}")
            items = []
            with open(txt_indice, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    parts = line.split("|")
                    if len(parts) == 2:
                        items.append((int(parts[0]), parts[1]))

            self.log_idx(f"📌 Se identificaron {len(items)} entradas en el índice:")
            for p, t in items:
                self.log_idx(f"    • Pág. {p}: {t}")

            pdf_indice_temp = os.path.join(os.path.dirname(pdf_salida), "temp_indice_page_ui.pdf")

            self.log_idx("\n📐 Maquetando preliminarmente para calcular desfase de páginas...")
            crear_pagina_indice_pdf(items, 0, pdf_indice_temp)
            
            doc_medida = pymupdf.open(pdf_indice_temp)
            desfase_real = doc_medida.page_count
            doc_medida.close()
            self.log_idx(f"ℹ️ El índice ocupará {desfase_real} página(s). Desfase aplicado a contenidos: +{desfase_real} pág(s).")

            self.log_idx("🎨 Generando páginas visuales definitivas del índice...")
            crear_pagina_indice_pdf(items, desfase_real, pdf_indice_temp)

            self.log_idx("\n🧩 Ensamblando índice con el documento original...")
            ensamblar_y_vincular_pdf(
                pdf_entrada, items, pdf_indice_temp, pdf_salida, desfase_real, log_func=self.log_idx
            )

            if os.path.exists(pdf_indice_temp):
                os.remove(pdf_indice_temp)
                self.log_idx("🧹 Archivos temporales eliminados.")

            tiempo_total = time.perf_counter() - tiempo_inicio
            self.log_idx("\n==================================================")
            self.log_idx("✅ ¡PROCESO FINALIZADO CON ÉXITO!")
            self.log_idx(f"⏱️ Tiempo total consumido: {tiempo_total:.2f} segundos")
            self.log_idx(f"📁 Archivo final disponible en:\n   {pdf_salida}")
            self.log_idx("==================================================\n")

            self.root.after(0, messagebox.showinfo, "Éxito", f"¡Índice y marcadores creados exitosamente!\n\nTiempo: {tiempo_total:.2f} s")

        except Exception as e:
            self.log_idx(f"\n❌ ERROR CRÍTICO DURANTE EL PROCESO: {str(e)}")
            self.root.after(0, messagebox.showerror, "Error Fatal", f"Ocurrió un error inesperado:\n{str(e)}")
        finally:
            self.is_processing_idx = False
            self.root.after(0, lambda: self.btn_process_idx.config(state="normal"))

    def _build_riesgo_demo_ui(self):
        f_card = ttk.LabelFrame(self.tab_riesgo_main, text=" Módulo de Análisis de Riesgo Fiscal y Muestreo ", padding=20)
        f_card.pack(fill="both", expand=True, padx=20, pady=20)

        lbl_title = ttk.Label(
            f_card, 
            text="🛡️ Módulo de Auditoría Preventiva y Detección de Anomalías", 
            font=("Arial", 12, "bold")
        )
        lbl_title.pack(anchor="w", pady=(0, 10))

        lbl_desc = ttk.Label(
            f_card, 
            text=(
                "Esta pestaña se encuentra en modo de presentación (Demo).\n\n"
                "Funcionalidades planificadas para este módulo:\n"
                " • Análisis de Ley de Benford para detección de alteraciones de cifras.\n"
                " • Algoritmos de muestreo estadístico (MUS, Estratificado, Aleatorio).\n"
                " • Identificación automática de pagos fraccionados o atípicos.\n"
                " • Generación de matriz de riesgos para dictamen fiscal."
            ),
            font=("Arial", 10),
            justify="left"
        )
        lbl_desc.pack(anchor="w", pady=5)


if __name__ == "__main__":
    if sys.platform == "win32":
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(2)
        except Exception:
            try:
                ctypes.windll.user32.SetProcessDPIAware()
            except Exception:
                pass

    root = tk.Tk()
    app = SuiteContableIntegrada(root)
    root.mainloop()
